# ============================================================
# MODULE ÉQUIPEMENTS
# Source : D:\Pompage\DATA_BASE_Merv.xlsx
# ============================================================

import openpyxl

EXCEL_PATH = r'D:\Pompage\DATA_BASE_Merv.xlsx'


def get_marques_panneaux():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb['Panneaux solaires']
        marques = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            marque = row[2]
            if marque and marque not in marques:
                marques.append(marque)
        wb.close()
        return sorted(marques)
    except Exception:
        return []


def get_modeles_panneaux(marque):
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Panneaux solaires']
    result = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[2] == marque:
            result.append({
                'marque':           row[2],
                'type':             row[3],
                'puissance_W':      row[4],
                'tension_V':        row[5],
                'Vmp_V':            row[6],
                'Imp_A':            row[7],
                'Voc_V':            row[8],
                'Isc_A':            row[9],
            })
    wb.close()
    return result


def get_marques_batteries():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb['Battéries']
        marques = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            marque = row[2]
            if marque and marque not in marques:
                marques.append(marque)
        wb.close()
        return sorted(marques)
    except Exception:
        return []


def get_modeles_batteries(marque):
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Battéries']
    result = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[2] == marque:
            rendement = row[6] / 100 if row[6] else 0.8
            dod = row[8] / 100 if row[8] else 0.8
            result.append({
                'marque':           row[2],
                'technologie':      row[3],
                'capacite_Ah':      row[4],
                'tension_V':        row[5],
                'rendement':        rendement,
                'temps_decharge_h': row[7],
                'dod':              dod,
            })
    wb.close()
    return result
